import asyncio
import os
import random
import string
import time
import json
import requests
from bs4 import BeautifulSoup
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
SPREADSHEET_ID = "1VUB2NdGSY0l3tuQAfkz8QV2XZpOj2khCB69r5zU1E5A"
SHEET_NAME = "CAPE CORAL FINAL"
RANGE_NAME = "R2:R"

user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
]

def authenticate_google_sheets():
    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    return build("sheets", "v4", credentials=creds)

def read_sheet_data(service):
    result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!{RANGE_NAME}"
    ).execute()
    return result.get("values", [])

def update_sheet_data(service, row, values):
    from openpyxl.utils import get_column_letter

    start_col = 20  # T
    end_col = start_col + len(values) - 1
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)

    range_name = f"{SHEET_NAME}!{start_letter}{row + 2}:{end_letter}{row + 2}"
    body = {"values": [values]}
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=range_name,
        valueInputOption="RAW",
        body=body,
    ).execute()

def extract_sitekey(content):
    soup = BeautifulSoup(content, "html.parser")
    turnstile_div = soup.find("div", {"class": "cf-turnstile"})
    if turnstile_div:
        return turnstile_div.get("data-sitekey")
    return None

async def inject_token(page, token, url):
    try:
        await page.evaluate(
            f"""
            () => {{
                const input = document.createElement('input');
                input.type = 'hidden';
                input.name = 'cf-turnstile-response';
                input.value = '{token}';
                document.querySelector('form')?.appendChild(input);
                document.querySelector('form')?.submit();
            }}
            """
        )
        return True
    except Exception as e:
        print(f"[!] Failed to inject CAPTCHA token: {e}")
        return False

async def solve_turnstile_captcha(api_key, sitekey, url):
    print(f"[+] Submitting CAPTCHA to 2Captcha...")
    response = requests.post(
        "http://2captcha.com/in.php",
        data={"key": api_key, "method": "turnstile", "sitekey": sitekey, "pageurl": url, "json": 1},
    )
    request_id = response.json().get("request")
    if not request_id:
        print(f"[!] CAPTCHA submission failed: {response.text}")
        return None

    print("[+] CAPTCHA submitted. Polling for result...")
    for _ in range(30):
        time.sleep(5)
        result = requests.get(
            "http://2captcha.com/res.php",
            params={"key": api_key, "action": "get", "id": request_id, "json": 1},
        ).json()
        if result.get("status") == 1:
            print("[+] CAPTCHA solved successfully.")
            return result.get("request")
        elif result.get("request") != "CAPCHA_NOT_READY":
            print(f"[!] CAPTCHA error: {result}")
            return None
    print("[!] CAPTCHA solving timed out.")
    return None

async def fetch_truepeoplesearch_data(url, api_key, browser):
    context = await browser.new_context(user_agent=random.choice(user_agents))
    page = await context.new_page()
    for attempt in range(3):
        try:
            await page.goto(url, timeout=60000)
            await page.wait_for_timeout(3000)
            content = await page.content()

            if "cf-turnstile" in content:
                sitekey = extract_sitekey(content)
                if not sitekey:
                    print("[!] Sitekey not found.")
                    return content

                token = await solve_turnstile_captcha(api_key, sitekey, url)
                if not token:
                    return content

                success = await inject_token(page, token, url)
                if success:
                    await page.wait_for_timeout(5000)
                    await page.reload(wait_until="networkidle")
                    await page.wait_for_timeout(3000)
                    content = await page.content()
                    return content
            else:
                return content
        except PlaywrightTimeout as e:
            print(f"[!] Timeout on attempt {attempt} for {url}: {e}")
        except Exception as e:
            print(f"[!] Error on attempt {attempt}: {e}")
    return None

def normalize_and_sort(name):
    return "".join(sorted(name.lower().replace(",", "").replace("jr", "").replace(".", "").strip()))

def match_entries(html_content, target_names):
    soup = BeautifulSoup(html_content, "html.parser")
    results = soup.find_all("div", class_="result-content")
    matches = []

    normalized_targets = [normalize_and_sort(name) for name in target_names]

    for result in results:
        name_tag = result.find("a", class_="h4")
        if name_tag:
            name_text = name_tag.text.strip()
            result_name_parts = [normalize_and_sort(part) for part in name_text.split()]
            for target in normalized_targets:
                if any(target in part for part in result_name_parts):
                    href = name_tag.get("href", "")
                    matches.append(f"{name_text} - https://www.truepeoplesearch.com{href}")
                    break
    return matches

async def main():
    service = authenticate_google_sheets()
    values = read_sheet_data(service)
    api_key = os.getenv("TWOCAPTCHA_API_KEY")
    if not api_key:
        print("[!] Missing 2Captcha API key.")
        return

    async with async_playwright() as p:
        browser = await p.firefox.launch(headless=True)
        for row, row_data in enumerate(values):
            if not row_data:
                continue
            url = row_data[0]
            print(f"[>] Processing row {row + 2}: {url}")

            html = await fetch_truepeoplesearch_data(url, api_key, browser)
            if not html:
                print(f"[!] Failed to fetch content for row {row + 2}")
                continue

            reference_names = [service.spreadsheets().values().get(
                spreadsheetId=SPREADSHEET_ID,
                range=f"{SHEET_NAME}!D{row + 2}:J{row + 2}"
            ).execute().get("values", [[]])[0]]

            flat_names = [name for sublist in reference_names for name in sublist]
            matches = match_entries(html, flat_names)

            if matches:
                update_sheet_data(service, row, matches)
                print(f"[✓] Row {row + 2} updated with {len(matches)} match(es).")
            else:
                print(f"[x] No matches found for row {row + 2}.")

        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())
